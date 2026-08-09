"""Bounded, chunked text extraction per format.

Every extractor reports how far it actually got: hitting a safety budget yields
``partial_due_to_safety_limit`` rather than a quiet truncation, and a document
we cannot read yields ``needs_ocr`` / ``needs_converter`` / ``unsupported_format``
rather than an empty success.
"""

from __future__ import annotations

import csv as _csv
import io
import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field

from origenlab_email_pipeline.chilecompra_anexo_evidence.archive import (
    ArchiveLimits,
    ArchiveSafetyError,
    preflight_archive,
    verify_actual_expansion,
)
from origenlab_email_pipeline.chilecompra_anexo_evidence.constants import (
    DEFAULT_ARCHIVE_VERIFY_CHUNK_BYTES,
    DEFAULT_CSV_ROWS_PER_CHUNK,
    DEFAULT_MAX_CHARS_PER_ATTACHMENT,
    DEFAULT_MAX_CHUNK_CHARS,
    DEFAULT_MAX_CSV_ROWS,
    DEFAULT_MAX_PDF_PAGES,
    DEFAULT_MAX_XLSX_CELLS,
    DEFAULT_MAX_XLSX_NONEMPTY_CELLS,
    DEFAULT_MAX_XML_CHARS,
    DEFAULT_PDF_MIN_PAGE_CHARS,
    DEFAULT_XLSX_ROWS_PER_CHUNK,
    FORMAT_CSV,
    FORMAT_DOC,
    FORMAT_DOCX,
    FORMAT_IMAGE,
    FORMAT_LEGACY_OLE,
    FORMAT_PDF,
    FORMAT_PPTX,
    FORMAT_RAR,
    FORMAT_SEVENZIP,
    FORMAT_TXT,
    FORMAT_XLS,
    FORMAT_XLSX,
    FORMAT_XML,
    OUTCOME_CORRUPT,
    OUTCOME_ENCRYPTED,
    OUTCOME_EXTRACTED_EMPTY,
    OUTCOME_EXTRACTION_FAILED,
    OUTCOME_EXTRACTION_SUCCESS,
    OUTCOME_NEEDS_CONVERTER,
    OUTCOME_NEEDS_OCR,
    OUTCOME_PARTIAL_DUE_TO_SAFETY_LIMIT,
    OUTCOME_UNSUPPORTED_FORMAT,
    REASON_EMBEDDED_CONTENT_UNPROCESSED,
    REASON_OOXML_CONTAINER_UNSAFE,
)

_WS_RE = re.compile(r"[ \t\r\f\v]+")


@dataclass(frozen=True)
class ExtractionLimits:
    max_pdf_pages: int = DEFAULT_MAX_PDF_PAGES
    max_chunk_chars: int = DEFAULT_MAX_CHUNK_CHARS
    max_chars_per_attachment: int = DEFAULT_MAX_CHARS_PER_ATTACHMENT
    max_csv_rows: int = DEFAULT_MAX_CSV_ROWS
    max_xlsx_cells: int = DEFAULT_MAX_XLSX_CELLS
    max_xlsx_nonempty_cells: int = DEFAULT_MAX_XLSX_NONEMPTY_CELLS
    max_xml_chars: int = DEFAULT_MAX_XML_CHARS
    csv_rows_per_chunk: int = DEFAULT_CSV_ROWS_PER_CHUNK
    xlsx_rows_per_chunk: int = DEFAULT_XLSX_ROWS_PER_CHUNK
    pdf_min_page_chars: int = DEFAULT_PDF_MIN_PAGE_CHARS


@dataclass(frozen=True)
class ChunkDraft:
    """Text plus its locator; IDs are assigned later by the bundle builder."""

    locator_type: str
    locator: dict[str, object]
    text: str


@dataclass
class ExtractionOutput:
    outcome: str
    extractor: str
    chunks: list[ChunkDraft] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    page_count: int | None = None
    sheet_count: int | None = None
    row_count: int | None = None
    error_message: str | None = None

    @property
    def char_count(self) -> int:
        return sum(len(c.text) for c in self.chunks)


def normalize_text(value: str) -> str:
    if not value:
        return ""
    value = value.replace("\x00", " ")
    value = _WS_RE.sub(" ", value)
    value = re.sub(r" *\n *", "\n", value)
    value = re.sub(r"\n\s*\n+", "\n\n", value)
    return value.strip()


def decode_text_bytes(payload: bytes) -> tuple[str, str]:
    """Decode with the encodings the portal actually emits; never raise."""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return payload.decode(encoding), encoding
        except (UnicodeDecodeError, LookupError):
            continue
    return payload.decode("utf-8", errors="replace"), "utf-8-replace"


def _split_text(text: str, limit: int) -> list[str]:
    if len(text) <= limit:
        return [text]
    return [text[i : i + limit] for i in range(0, len(text), limit)]


class _Budget:
    """Character budget shared across one attachment's chunks."""

    def __init__(self, limit: int) -> None:
        self.limit = limit
        self.used = 0
        self.exhausted = False

    def take(self, text: str) -> str:
        if self.exhausted:
            return ""
        remaining = self.limit - self.used
        if remaining <= 0:
            self.exhausted = True
            return ""
        if len(text) > remaining:
            self.exhausted = True
            self.used = self.limit
            return text[:remaining]
        self.used += len(text)
        return text


def _finalize(
    output: ExtractionOutput,
    *,
    budget: _Budget,
    hit_limit: bool,
) -> ExtractionOutput:
    if output.outcome not in {"", OUTCOME_EXTRACTION_SUCCESS}:
        return output
    if hit_limit or budget.exhausted:
        output.outcome = OUTCOME_PARTIAL_DUE_TO_SAFETY_LIMIT
        return output
    output.outcome = (
        OUTCOME_EXTRACTION_SUCCESS if output.chunks else OUTCOME_EXTRACTED_EMPTY
    )
    return output


# --- PDF ---------------------------------------------------------------------


def extract_pdf(payload: bytes, *, limits: ExtractionLimits) -> ExtractionOutput:
    out = ExtractionOutput(outcome="", extractor="pymupdf_page_text")
    try:
        import fitz  # type: ignore
    except ImportError:
        out.outcome = OUTCOME_NEEDS_CONVERTER
        out.warnings.append("dependency_unavailable:pymupdf")
        return out

    try:
        document = fitz.open(stream=payload, filetype="pdf")
    except Exception as exc:  # noqa: BLE001 - any parse failure means corrupt input
        out.outcome = OUTCOME_CORRUPT
        out.error_message = str(exc)[:300]
        return out

    try:
        if getattr(document, "needs_pass", False):
            out.outcome = OUTCOME_ENCRYPTED
            out.warnings.append("pdf_password_protected")
            return out
        page_count = int(document.page_count)
        out.page_count = page_count
        budget = _Budget(limits.max_chars_per_attachment)
        hit_limit = page_count > limits.max_pdf_pages
        if hit_limit:
            out.warnings.append(
                f"pdf_page_limit:{limits.max_pdf_pages}/{page_count}"
            )
        pages_with_text = 0
        for index in range(min(page_count, limits.max_pdf_pages)):
            try:
                page_text = document.load_page(index).get_text("text") or ""
            except Exception as exc:  # noqa: BLE001 - one bad page must not kill the doc
                out.warnings.append(f"pdf_page_unreadable:{index + 1}:{type(exc).__name__}")
                continue
            text = normalize_text(page_text)
            if len(text.strip()) >= limits.pdf_min_page_chars:
                pages_with_text += 1
            if not text:
                continue
            for part_index, part in enumerate(_split_text(text, limits.max_chunk_chars)):
                allowed = budget.take(part)
                if not allowed:
                    break
                out.chunks.append(
                    ChunkDraft(
                        locator_type="pdf_page",
                        locator={
                            "page": index + 1,
                            "page_count": page_count,
                            "part": part_index + 1,
                        },
                        text=allowed,
                    )
                )
            if budget.exhausted:
                out.warnings.append("char_budget_reached")
                break
        if pages_with_text == 0 and page_count > 0:
            out.outcome = OUTCOME_NEEDS_OCR
            out.warnings.append("pdf_no_extractable_text")
            return out
        if 0 < pages_with_text < min(page_count, limits.max_pdf_pages):
            out.warnings.append(
                f"pdf_pages_without_text:{min(page_count, limits.max_pdf_pages) - pages_with_text}"
            )
        return _finalize(out, budget=budget, hit_limit=hit_limit)
    finally:
        try:
            document.close()
        except Exception:  # noqa: BLE001 - close failures are not extraction failures
            pass


# --- OOXML container safety --------------------------------------------------


def ooxml_container_guard(
    payload: bytes,
    *,
    archive_limits: ArchiveLimits,
    extractor: str,
) -> ExtractionOutput | None:
    """Apply the generic-ZIP safety envelope before any OOXML part is decompressed.

    DOCX/XLSX/PPTX are ZIP containers, so without this a malicious OOXML file
    would reach ``zipfile``/``openpyxl`` directly and decompress unbounded bytes.
    Runs in two stages: the declared central-directory preflight, then bounded
    verification of what the members *actually* expand to, because a lying
    header passes the first stage and the parsers issue their own unbounded
    reads. Returns an explicit incomplete outcome, or ``None`` to proceed.
    """
    try:
        preflight = preflight_archive(payload, limits=archive_limits)
    except ArchiveSafetyError as exc:
        return ExtractionOutput(
            outcome=OUTCOME_CORRUPT,
            extractor=extractor,
            warnings=["ooxml_container_unreadable"],
            error_message=str(exc)[:300],
        )
    if preflight.encrypted_member_count:
        return ExtractionOutput(
            outcome=OUTCOME_ENCRYPTED,
            extractor=extractor,
            warnings=[
                REASON_OOXML_CONTAINER_UNSAFE,
                f"encrypted_members:{preflight.encrypted_member_count}",
            ],
        )
    if preflight.rejected:
        return ExtractionOutput(
            outcome=OUTCOME_PARTIAL_DUE_TO_SAFETY_LIMIT,
            extractor=extractor,
            warnings=[REASON_OOXML_CONTAINER_UNSAFE, *preflight.reason_codes],
        )

    expansion = verify_actual_expansion(payload, limits=archive_limits)
    if expansion.rejected:
        return ExtractionOutput(
            outcome=OUTCOME_PARTIAL_DUE_TO_SAFETY_LIMIT,
            extractor=extractor,
            warnings=[REASON_OOXML_CONTAINER_UNSAFE, *expansion.reason_codes],
            error_message=(
                f"actual expansion rejected at member {expansion.offending_member!r}"
                if expansion.offending_member
                else None
            ),
        )
    return None


def _read_member_capped(
    archive: zipfile.ZipFile,
    name: str,
    *,
    max_bytes: int,
    chunk_bytes: int = DEFAULT_ARCHIVE_VERIFY_CHUNK_BYTES,
) -> tuple[bytes, bool]:
    """Read one part through a cap so an understated header cannot overrun it.

    Accumulates fixed-size chunks rather than issuing one large read: a single
    ``read(max_bytes + 1)`` lets zipfile decompress greedily and spike well past
    the part's real size before the cap is applied.
    """
    collected: list[bytes] = []
    total = 0
    with archive.open(name) as handle:
        while total <= max_bytes:
            block = handle.read(min(chunk_bytes, max_bytes + 1 - total))
            if not block:
                break
            collected.append(block)
            total += len(block)
    data = b"".join(collected)
    if len(data) > max_bytes:
        return data[:max_bytes], True
    return data, False


# --- DOCX --------------------------------------------------------------------

_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _docx_node_text(node: ET.Element) -> str:
    parts = [t.text or "" for t in node.iter(f"{_W_NS}t")]
    return normalize_text("".join(parts))


def _docx_part_chunks(
    xml_bytes: bytes,
    *,
    section: str,
    limits: ExtractionLimits,
    budget: _Budget,
    warnings: list[str],
) -> list[ChunkDraft]:
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as exc:
        warnings.append(f"docx_part_unparsable:{section}:{type(exc).__name__}")
        return []
    chunks: list[ChunkDraft] = []
    body = root.find(f"{_W_NS}body")
    container = body if body is not None else root
    # Content inside a table is emitted once, with its table; nested tables are
    # rendered by their parent rather than a second time at top level.
    tables = list(container.iter(f"{_W_NS}tbl"))
    nested_paragraph_ids = {
        id(node) for table in tables for node in table.iter(f"{_W_NS}p")
    }
    nested_table_ids = {
        id(node)
        for table in tables
        for node in table.iter(f"{_W_NS}tbl")
        if node is not table
    }
    paragraph_index = 0
    table_index = 0
    for node in container.iter():
        tag = node.tag
        if tag == f"{_W_NS}p":
            if id(node) in nested_paragraph_ids:
                continue
            text = _docx_node_text(node)
            paragraph_index += 1
            if not text:
                continue
            for part in _split_text(text, limits.max_chunk_chars):
                allowed = budget.take(part)
                if not allowed:
                    return chunks
                chunks.append(
                    ChunkDraft(
                        locator_type="docx_paragraph",
                        locator={"section": section, "paragraph": paragraph_index},
                        text=allowed,
                    )
                )
        elif tag == f"{_W_NS}tbl":
            if id(node) in nested_table_ids:
                continue
            table_index += 1
            rows: list[str] = []
            for row_index, row in enumerate(node.findall(f"{_W_NS}tr"), start=1):
                cells = [_docx_node_text(cell) for cell in row.findall(f"{_W_NS}tc")]
                if any(cells):
                    rows.append(f"[r{row_index}] " + " | ".join(cells))
            if not rows:
                continue
            for part in _split_text("\n".join(rows), limits.max_chunk_chars):
                allowed = budget.take(part)
                if not allowed:
                    return chunks
                chunks.append(
                    ChunkDraft(
                        locator_type="docx_table",
                        locator={"section": section, "table": table_index},
                        text=allowed,
                    )
                )
    return chunks


def extract_docx(
    payload: bytes,
    *,
    limits: ExtractionLimits,
    archive_limits: ArchiveLimits | None = None,
) -> ExtractionOutput:
    """Read body paragraphs, tables, headers, and footers (specs hide in all four)."""
    active_archive = archive_limits or ArchiveLimits()
    blocked = ooxml_container_guard(
        payload, archive_limits=active_archive, extractor="ooxml_docx"
    )
    if blocked is not None:
        return blocked

    out = ExtractionOutput(outcome="", extractor="ooxml_docx")
    budget = _Budget(limits.max_chars_per_attachment)
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            names = set(archive.namelist())
            if any(n.startswith("word/embeddings/") for n in names):
                out.warnings.append(REASON_EMBEDDED_CONTENT_UNPROCESSED)
            if any(n.startswith("word/vbaProject") or n.endswith(".bin") for n in names):
                out.warnings.append("macro_or_binary_part_present_not_executed")
            ordered_parts = ["word/document.xml"]
            ordered_parts += sorted(n for n in names if re.fullmatch(r"word/header\d*\.xml", n))
            ordered_parts += sorted(n for n in names if re.fullmatch(r"word/footer\d*\.xml", n))
            if "word/document.xml" not in names:
                out.outcome = OUTCOME_CORRUPT
                out.error_message = "missing word/document.xml"
                return out
            for part in ordered_parts:
                if part not in names:
                    continue
                section = (
                    "body"
                    if part == "word/document.xml"
                    else part.removeprefix("word/").removesuffix(".xml")
                )
                data, truncated = _read_member_capped(
                    archive,
                    part,
                    max_bytes=active_archive.max_member_uncompressed_bytes,
                )
                if truncated:
                    out.warnings.append(f"ooxml_part_truncated:{section}")
                out.chunks.extend(
                    _docx_part_chunks(
                        data,
                        section=section,
                        limits=limits,
                        budget=budget,
                        warnings=out.warnings,
                    )
                )
                if budget.exhausted:
                    out.warnings.append("char_budget_reached")
                    break
    except zipfile.BadZipFile as exc:
        out.outcome = OUTCOME_CORRUPT
        out.error_message = str(exc)[:300]
        return out
    except Exception as exc:  # noqa: BLE001 - surface as failure, never as silence
        out.outcome = OUTCOME_EXTRACTION_FAILED
        out.error_message = str(exc)[:300]
        return out

    result = _finalize(
        out,
        budget=budget,
        hit_limit=any(w.startswith("ooxml_part_truncated:") for w in out.warnings),
    )
    if REASON_EMBEDDED_CONTENT_UNPROCESSED in out.warnings and result.outcome in {
        OUTCOME_EXTRACTION_SUCCESS,
        OUTCOME_EXTRACTED_EMPTY,
    }:
        result.outcome = OUTCOME_PARTIAL_DUE_TO_SAFETY_LIMIT
    return result


# --- XLSX --------------------------------------------------------------------


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def extract_xlsx(
    payload: bytes,
    *,
    limits: ExtractionLimits,
    archive_limits: ArchiveLimits | None = None,
) -> ExtractionOutput:
    """Walk every worksheet including hidden ones; no sheet/row/column sampling."""
    blocked = ooxml_container_guard(
        payload,
        archive_limits=archive_limits or ArchiveLimits(),
        extractor="openpyxl_all_sheets",
    )
    if blocked is not None:
        return blocked

    out = ExtractionOutput(outcome="", extractor="openpyxl_all_sheets")
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        out.outcome = OUTCOME_NEEDS_CONVERTER
        out.warnings.append("dependency_unavailable:openpyxl")
        return out

    budget = _Budget(limits.max_chars_per_attachment)
    hit_limit = False
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - malformed workbook
        out.outcome = OUTCOME_CORRUPT
        out.error_message = str(exc)[:300]
        return out

    try:
        sheet_names = list(workbook.sheetnames or [])
        out.sheet_count = len(sheet_names)
        cells_seen = 0
        nonempty_seen = 0
        for sheet_name in sheet_names:
            try:
                sheet = workbook[sheet_name]
            except Exception as exc:  # noqa: BLE001 - skip but record
                out.warnings.append(f"sheet_unreadable:{type(exc).__name__}")
                continue
            state = getattr(sheet, "sheet_state", "visible")
            if state != "visible":
                out.warnings.append(f"hidden_sheet_included:{sheet_name}")
            pending: list[str] = []
            pending_start = 1
            row_index = 0
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [_cell_to_text(v) for v in row]
                cells_seen += len(cells)
                nonempty = [c for c in cells if c]
                nonempty_seen += len(nonempty)
                if nonempty:
                    while cells and not cells[-1]:
                        cells.pop()
                    pending.append(f"[r{row_index}] " + " | ".join(cells))
                if cells_seen > limits.max_xlsx_cells or (
                    nonempty_seen > limits.max_xlsx_nonempty_cells
                ):
                    hit_limit = True
                    out.warnings.append("xlsx_cell_limit_reached")
                    break
                if len(pending) >= limits.xlsx_rows_per_chunk:
                    _flush_rows(
                        out,
                        budget,
                        pending,
                        sheet_name=sheet_name,
                        start_row=pending_start,
                        end_row=row_index,
                        limits=limits,
                    )
                    pending = []
                    pending_start = row_index + 1
                    if budget.exhausted:
                        break
            if pending and not budget.exhausted:
                _flush_rows(
                    out,
                    budget,
                    pending,
                    sheet_name=sheet_name,
                    start_row=pending_start,
                    end_row=max(row_index, pending_start),
                    limits=limits,
                )
            if hit_limit or budget.exhausted:
                break
        out.row_count = None
        return _finalize(out, budget=budget, hit_limit=hit_limit)
    except Exception as exc:  # noqa: BLE001
        out.outcome = OUTCOME_EXTRACTION_FAILED
        out.error_message = str(exc)[:300]
        return out
    finally:
        try:
            workbook.close()
        except Exception:  # noqa: BLE001
            pass


def _flush_rows(
    out: ExtractionOutput,
    budget: _Budget,
    rows: list[str],
    *,
    sheet_name: str,
    start_row: int,
    end_row: int,
    limits: ExtractionLimits,
) -> None:
    text = "\n".join(rows)
    if not text.strip():
        return
    for part in _split_text(text, limits.max_chunk_chars):
        allowed = budget.take(part)
        if not allowed:
            return
        out.chunks.append(
            ChunkDraft(
                locator_type="xlsx_rows",
                locator={
                    "sheet": sheet_name,
                    "start_row": start_row,
                    "end_row": end_row,
                },
                text=allowed,
            )
        )


# --- Legacy XLS --------------------------------------------------------------


def extract_xls(payload: bytes, *, limits: ExtractionLimits) -> ExtractionOutput:
    """Use xlrd when the data-tools group is installed; otherwise say so explicitly."""
    out = ExtractionOutput(outcome="", extractor="xlrd_legacy_xls")
    try:
        import xlrd  # type: ignore
    except ImportError:
        out.outcome = OUTCOME_NEEDS_CONVERTER
        out.warnings.append("dependency_unavailable:xlrd")
        return out

    budget = _Budget(limits.max_chars_per_attachment)
    try:
        book = xlrd.open_workbook(file_contents=payload)
    except Exception as exc:  # noqa: BLE001
        out.outcome = OUTCOME_CORRUPT
        out.error_message = str(exc)[:300]
        return out

    hit_limit = False
    cells_seen = 0
    out.sheet_count = book.nsheets
    for sheet_index in range(book.nsheets):
        sheet = book.sheet_by_index(sheet_index)
        pending: list[str] = []
        pending_start = 1
        row_index = 0
        for row_index in range(1, sheet.nrows + 1):
            values = sheet.row_values(row_index - 1)
            cells = [_cell_to_text(v) for v in values]
            cells_seen += len(cells)
            if any(cells):
                pending.append(f"[r{row_index}] " + " | ".join(cells))
            if cells_seen > limits.max_xlsx_cells:
                hit_limit = True
                out.warnings.append("xls_cell_limit_reached")
                break
            if len(pending) >= limits.xlsx_rows_per_chunk:
                _flush_rows(
                    out,
                    budget,
                    pending,
                    sheet_name=sheet.name,
                    start_row=pending_start,
                    end_row=row_index,
                    limits=limits,
                )
                pending = []
                pending_start = row_index + 1
                if budget.exhausted:
                    break
        if pending and not budget.exhausted:
            _flush_rows(
                out,
                budget,
                pending,
                sheet_name=sheet.name,
                start_row=pending_start,
                end_row=max(row_index, pending_start),
                limits=limits,
            )
        if hit_limit or budget.exhausted:
            break
    return _finalize(out, budget=budget, hit_limit=hit_limit)


# --- CSV / TXT / XML ---------------------------------------------------------


def extract_csv(payload: bytes, *, limits: ExtractionLimits) -> ExtractionOutput:
    out = ExtractionOutput(outcome="", extractor="csv_rows")
    text, encoding = decode_text_bytes(payload)
    if encoding != "utf-8":
        out.warnings.append(f"decoded_as:{encoding}")
    budget = _Budget(limits.max_chars_per_attachment)
    hit_limit = False
    pending: list[str] = []
    pending_start = 1
    row_index = 0
    try:
        reader = _csv.reader(io.StringIO(text))
        for row_index, row in enumerate(reader, start=1):
            if row_index > limits.max_csv_rows:
                hit_limit = True
                out.warnings.append(f"csv_row_limit:{limits.max_csv_rows}")
                break
            cells = [c.replace("\n", " ").strip() for c in row]
            if any(cells):
                pending.append(", ".join(cells))
            if len(pending) >= limits.csv_rows_per_chunk:
                _flush_csv(out, budget, pending, pending_start, row_index, limits)
                pending = []
                pending_start = row_index + 1
                if budget.exhausted:
                    out.warnings.append("char_budget_reached")
                    break
    except _csv.Error as exc:
        out.outcome = OUTCOME_EXTRACTION_FAILED
        out.error_message = str(exc)[:300]
        return out
    if pending and not budget.exhausted:
        _flush_csv(out, budget, pending, pending_start, max(row_index, pending_start), limits)
    out.row_count = row_index
    return _finalize(out, budget=budget, hit_limit=hit_limit)


def _flush_csv(
    out: ExtractionOutput,
    budget: _Budget,
    rows: list[str],
    start_row: int,
    end_row: int,
    limits: ExtractionLimits,
) -> None:
    text = "\n".join(rows)
    if not text.strip():
        return
    for part in _split_text(text, limits.max_chunk_chars):
        allowed = budget.take(part)
        if not allowed:
            return
        out.chunks.append(
            ChunkDraft(
                locator_type="csv_rows",
                locator={"start_row": start_row, "end_row": end_row},
                text=allowed,
            )
        )


def extract_txt(payload: bytes, *, limits: ExtractionLimits) -> ExtractionOutput:
    out = ExtractionOutput(outcome="", extractor="plain_text")
    text, encoding = decode_text_bytes(payload)
    if encoding != "utf-8":
        out.warnings.append(f"decoded_as:{encoding}")
    normalized = normalize_text(text)
    budget = _Budget(limits.max_chars_per_attachment)
    for index, part in enumerate(_split_text(normalized, limits.max_chunk_chars), start=1):
        allowed = budget.take(part)
        if not allowed:
            out.warnings.append("char_budget_reached")
            break
        out.chunks.append(
            ChunkDraft(
                locator_type="text_span",
                locator={"part": index},
                text=allowed,
            )
        )
    return _finalize(out, budget=budget, hit_limit=False)


def extract_xml(payload: bytes, *, limits: ExtractionLimits) -> ExtractionOutput:
    """Parse locally only: entity declarations are refused, never resolved."""
    out = ExtractionOutput(outcome="", extractor="xml_text_nodes")
    if b"<!ENTITY" in payload[: 64 * 1024].upper():
        out.outcome = OUTCOME_EXTRACTION_FAILED
        out.error_message = "xml entity declaration rejected"
        out.warnings.append("xml_entity_declaration_rejected")
        return out
    if len(payload) > limits.max_xml_chars:
        out.warnings.append(f"xml_size_limit:{limits.max_xml_chars}")
        payload = payload[: limits.max_xml_chars]
        hit_limit = True
    else:
        hit_limit = False
    text, encoding = decode_text_bytes(payload)
    if encoding != "utf-8":
        out.warnings.append(f"decoded_as:{encoding}")
    try:
        root = ET.fromstring(text.replace("\x00", ""))
    except ET.ParseError as exc:
        out.outcome = OUTCOME_EXTRACTION_FAILED
        out.error_message = str(exc)[:300]
        return out
    budget = _Budget(limits.max_chars_per_attachment)
    values: list[str] = []
    for element in root.iter():
        for raw in (element.text, element.tail):
            if raw and raw.strip():
                values.append(raw.strip())
    combined = normalize_text("\n".join(values))
    for index, part in enumerate(_split_text(combined, limits.max_chunk_chars), start=1):
        allowed = budget.take(part)
        if not allowed:
            out.warnings.append("char_budget_reached")
            break
        out.chunks.append(
            ChunkDraft(
                locator_type="xml_text",
                locator={"part": index},
                text=allowed,
            )
        )
    return _finalize(out, budget=budget, hit_limit=hit_limit)


# --- Dispatch ----------------------------------------------------------------

_NEEDS_CONVERTER_FORMATS = {FORMAT_DOC, FORMAT_LEGACY_OLE, FORMAT_PPTX}
_UNSUPPORTED_FORMATS = {FORMAT_RAR, FORMAT_SEVENZIP}


def extract_payload(
    payload: bytes,
    *,
    detected_format: str,
    limits: ExtractionLimits | None = None,
    archive_limits: ArchiveLimits | None = None,
) -> ExtractionOutput:
    """Route to the right bounded extractor; unknown formats are named, not skipped."""
    active = limits or ExtractionLimits()
    active_archive = archive_limits or ArchiveLimits()
    if detected_format == FORMAT_PDF:
        return extract_pdf(payload, limits=active)
    if detected_format == FORMAT_DOCX:
        return extract_docx(payload, limits=active, archive_limits=active_archive)
    if detected_format == FORMAT_XLSX:
        return extract_xlsx(payload, limits=active, archive_limits=active_archive)
    if detected_format == FORMAT_XLS:
        return extract_xls(payload, limits=active)
    if detected_format == FORMAT_CSV:
        return extract_csv(payload, limits=active)
    if detected_format == FORMAT_TXT:
        return extract_txt(payload, limits=active)
    if detected_format == FORMAT_XML:
        return extract_xml(payload, limits=active)
    if detected_format == FORMAT_IMAGE:
        return ExtractionOutput(
            outcome=OUTCOME_NEEDS_OCR,
            extractor="none",
            warnings=["image_attachment_requires_ocr"],
        )
    if detected_format in _NEEDS_CONVERTER_FORMATS:
        return ExtractionOutput(
            outcome=OUTCOME_NEEDS_CONVERTER,
            extractor="none",
            warnings=[f"no_safe_converter_for:{detected_format}"],
        )
    if detected_format in _UNSUPPORTED_FORMATS:
        return ExtractionOutput(
            outcome=OUTCOME_UNSUPPORTED_FORMAT,
            extractor="none",
            warnings=[f"unsupported_archive_format:{detected_format}"],
        )
    return ExtractionOutput(
        outcome=OUTCOME_UNSUPPORTED_FORMAT,
        extractor="none",
        warnings=[f"unsupported_format:{detected_format}"],
    )
